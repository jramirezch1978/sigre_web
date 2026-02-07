# -*- coding: utf-8 -*-
"""
Script para extraer todos los datos del archivo Excel 
'Inventario de composición de módulo.xlsx'
incluyendo hipervínculos (URLs) de los documentos de Google Docs.
"""

import openpyxl
import os
import json

path = r'e:\Work\sigre_web\restaurante.pe'
for f in os.listdir(path):
    if f.endswith('.xlsx'):
        filepath = os.path.join(path, f)
        break

wb = openpyxl.load_workbook(filepath)

output = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Recopilar hipervínculos del sheet
    hyperlinks_map = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:
                hyperlinks_map[cell.coordinate] = cell.hyperlink.target
    
    # Encontrar filas con datos reales (no vacías)
    sheet_data = {
        "name": sheet_name,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "hyperlinks_count": len(hyperlinks_map),
        "hyperlinks": hyperlinks_map,
        "rows": []
    }
    
    empty_count = 0
    for row in ws.iter_rows(min_row=1, values_only=False):
        row_values = []
        row_has_data = False
        for cell in row:
            val = cell.value
            if val is not None:
                row_has_data = True
            row_values.append(str(val) if val is not None else "")
        
        if row_has_data:
            empty_count = 0
            sheet_data["rows"].append(row_values)
        else:
            empty_count += 1
            if empty_count > 5:
                # Más de 5 filas vacías seguidas, dejamos de leer
                continue
    
    output[sheet_name] = sheet_data

# Escribir resultado
output_file = os.path.join(path, "extracted_data.json")
with open(output_file, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

# Imprimir resumen
print("=" * 80)
print("RESUMEN DE EXTRACCIÓN")
print("=" * 80)
for sheet_name, data in output.items():
    print(f"\nHOJA: {data['name']}")
    print(f"   Filas con datos: {len(data['rows'])}")
    print(f"   Hipervinculos encontrados: {data['hyperlinks_count']}")
    if data['hyperlinks']:
        print(f"   URLs:")
        for coord, url in data['hyperlinks'].items():
            print(f"     [{coord}] -> {url}")

print(f"\n\nDatos completos guardados en: {output_file}")
